from openpyxl import Workbook
from openpyxl import load_workbook as lwb
from openpyxl.utils.dataframe import dataframe_to_rows
import os

path_desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')

def excel_Convert(name_archivo,df):

    wb = Workbook()
    filesheet = f"{path_desktop}/{name_archivo}.xlsx"
    ws= wb.active
    for r in dataframe_to_rows(df,index=True, header=True):
          ws.append(r)
    wb.save(filesheet)
    pass

def formatoExcel(name_archivo, nombre_hoja):
     wb = lwb(name_archivo)
     hoja = wb[nombre_hoja]
     min_col = wb.active.min_column
     max_col = wb.active.max_column
     min_fila = wb.active.min_row
     max_fila = wb.active.max_row
  